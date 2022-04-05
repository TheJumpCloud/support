import requests
from openpyxl import Workbook

def get_user_data():
    try:
        data_exists = True
        skip_count = 0
        users_data = []
        while data_exists:
            url = "https://console.jumpcloud.com/api/systemusers"
            querystring = {"limit":"100","skip":f"{skip_count}","sort":"email"}
            response = requests.request("GET", url, headers=HEADERS, params=querystring)
            if response.json()['results']:
                data_exists = True
            else:
                data_exists = False
            users_data.extend(response.json()['results'])
            skip_count += 100
        return users_data
    except Exception as e:
        print(f"Exception occurred while fetching the user data : {str(e)}")
        return []

def get_system_details(systemid):
    try:
        url = f"https://console.jumpcloud.com/api/systems/{systemid}"
        response = requests.request("GET", url, headers=HEADERS)
        if response.json():
            return response.json()
    except Exception as e:
        print(f"Exception occurred while fetching the system data : {str(e)}")
        return None

def get_system_association(user):
    try:
        data_exists = True
        device_access = []
        url = f"https://console.jumpcloud.com/api/v2/users/{user}/systems"
        response = requests.request("GET", url, headers=HEADERS)
        if response.json():
            device_list = response.json()
            for device in device_list:
                sys_info = get_system_details(systemid=device['id'])
                temp = {}
                temp['os'] = sys_info['os']
                temp['id'] = device['id']
                temp['device_name'] = sys_info['displayName']
                if "serialNumber" in sys_info:
                    temp['serial_number'] = sys_info['serialNumber']
                else:
                    temp['serial_number'] = "NA"
                if not device['compiledAttributes']:
                    temp['user_access'] = "Normal User"
                else:
                    # print(device['compiledAttributes'])
                    temp['user_access'] = "Administrator" if device['compiledAttributes']['sudo']['enabled'] else "Normal User"
                device_access.append(temp)
        return device_access
    except Exception as e:
        print(f"Exception occurred while fetching the user and system mapping data : {str(e)}")
        return []

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Process some integers.')
    parser.add_argument('--apikey', help='JumpCloud API Key')
    parser.add_argument('--output_file', help='Output Excel filepath')

    args = parser.parse_args()
    HEADERS = {"x-api-key": args.apikey}
    user_list = get_user_data()
    user_col_id_name_mapping = {
        1 : "email",
        2 : "state",
        3 : "totp_enabled"
    }
    device_col_id_name_mapping = {
        4 : "os",
        5 : "device_name",
        6 : "serial_number",
        7 : "user_access"
    }

    wb = Workbook()
    ws = wb.create_sheet("jc_user_access")
    row_number, column_number = 1,1
    for u_col, u_val in user_col_id_name_mapping.items():
        ws.cell(row_number,u_col).value = u_val
    for d_col, d_val in device_col_id_name_mapping.items():
        ws.cell(row_number,d_col).value = d_val
    row_number += 1
    for user in user_list:
        print(f"Collecting details for User: {user['email']}")
        device_access = get_system_association(user=user['id'])
        for u_col, u_val in user_col_id_name_mapping.items():
            ws.cell(row_number,u_col).value = user[u_val]
        for loop_id, device in enumerate(device_access):
            if loop_id > 0:
                for u_col, u_val in user_col_id_name_mapping.items():
                    ws.cell(row_number,u_col).value = user[u_val]                
            for d_col, d_val in device_col_id_name_mapping.items():
                ws.cell(row_number,d_col).value = device[d_val]            
            row_number += 1
    wb.active = wb['jc_user_access']
    wb.save(filename=args.output_file)